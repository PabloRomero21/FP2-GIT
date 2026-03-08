import csv
# Importamos las clases necesarias de nuestros otros módulos
from registro import RegistroClasificacion, RegistroRegresion
from dataset import DataSetClasificacion, DataSetRegresion

# Intentamos importar openpyxl para los archivos de Excel
try:
    from openpyxl import load_workbook
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

class FactoriaCSV:
    """Clase factoría encargada de construir DataSets a partir de archivos CSV."""

    @staticmethod
    def crear_dataset_clasificacion(ruta_fichero, indice_objetivo=-1):
        """Crea y devuelve un objeto DataSetClasificacion leyendo un CSV."""
        dataset = DataSetClasificacion()
        
        with open(ruta_fichero, mode='r', encoding='utf-8') as archivo:
            lector_csv = csv.reader(archivo)
            # 1. Leer la primera fila (Cabeceras)
            cabeceras = next(lector_csv)
            # Extraemos el nombre de la columna objetivo y dejamos el resto como atributos
            cabeceras.pop(indice_objetivo) 
            dataset.nombres_atributos = cabeceras
            
            # 2. Leer el resto de las filas (Datos)
            for fila in lector_csv:
                if not fila: continue # Evitar líneas en blanco
                
                # Extraemos el valor objetivo basándonos en el índice
                objetivo = fila.pop(indice_objetivo)
                # El resto de la lista 'fila' son los atributos numéricos
                registro = RegistroClasificacion(fila, objetivo)
                
                # Aprovechamos el método que creamos en la Etapa 3
                dataset.agregar_registro(registro)
                
        return dataset

    @staticmethod
    def crear_dataset_regresion(ruta_fichero, indice_objetivo=-1):
        """Crea y devuelve un objeto DataSetRegresion leyendo un CSV."""
        dataset = DataSetRegresion()
        
        with open(ruta_fichero, mode='r', encoding='utf-8') as archivo:
            lector_csv = csv.reader(archivo)
            # 1. Leer cabeceras
            cabeceras = next(lector_csv)
            cabeceras.pop(indice_objetivo)
            dataset.nombres_atributos = cabeceras
            
            # 2. Leer datos
            for fila in lector_csv:
                if not fila: continue
                
                objetivo = fila.pop(indice_objetivo)
                registro = RegistroRegresion(fila, objetivo)
                dataset.agregar_registro(registro)
                
        return dataset


class FactoriaXLS:
    """Clase factoría encargada de construir DataSets a partir de archivos de Excel (.xlsx)."""

    @staticmethod
    def _leer_datos_excel(ruta_fichero):
        """Método auxiliar oculto para extraer filas de un Excel genérico."""
        if not EXCEL_DISPONIBLE:
            raise ImportError("Debe instalar 'openpyxl' para leer archivos Excel (pip install openpyxl).")
        
        wb = load_workbook(filename=ruta_fichero, data_only=True)
        hoja = wb.active
        # Convertimos las filas (tuplas) en listas estándar de Python
        return [[celda.value for celda in fila] for fila in hoja.iter_rows()]

    @staticmethod
    def crear_dataset_clasificacion(ruta_fichero, indice_objetivo=-1):
        """Crea y devuelve un objeto DataSetClasificacion leyendo un Excel."""
        dataset = DataSetClasificacion()
        filas = FactoriaXLS._leer_datos_excel(ruta_fichero)
        
        if filas:
            cabeceras = filas[0]
            cabeceras.pop(indice_objetivo)
            dataset.nombres_atributos = cabeceras
            
            for fila in filas[1:]:
                # Asegurarnos de que no sea una fila vacía
                if any(celda is not None for celda in fila): 
                    objetivo = fila.pop(indice_objetivo)
                    registro = RegistroClasificacion(fila, objetivo)
                    dataset.agregar_registro(registro)
                    
        return dataset

    @staticmethod
    def crear_dataset_regresion(ruta_fichero, indice_objetivo=-1):
        """Crea y devuelve un objeto DataSetRegresion leyendo un Excel."""
        dataset = DataSetRegresion()
        filas = FactoriaXLS._leer_datos_excel(ruta_fichero)
        
        if filas:
            cabeceras = filas[0]
            cabeceras.pop(indice_objetivo)
            dataset.nombres_atributos = cabeceras
            
            for fila in filas[1:]:
                if any(celda is not None for celda in fila):
                    objetivo = fila.pop(indice_objetivo)
                    registro = RegistroRegresion(fila, objetivo)
                    dataset.agregar_registro(registro)
                    
        return dataset
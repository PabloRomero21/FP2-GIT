import openpyxl
from proyectos import Proyecto, ProyectoConcedido, ProyectoContrato
from gestor import Gestor_Proyectos, Gestor_ProyectosConcedidos, Gestor_ProyectosContrato

class Factoria:
    """
    Clase encargada de leer los ficheros Excel (.xlsx) nativos y construir 
    los objetos del modelo, distribuyéndolos en sus gestores correspondientes.
    """
    
    def __init__(self, ruta_anexo1: str, ruta_anexo2: str, ruta_anexo3: str, ruta_anexo4: str):
        self.ruta_anexo1 = ruta_anexo1
        self.ruta_anexo2 = ruta_anexo2
        self.ruta_anexo3 = ruta_anexo3
        self.ruta_anexo4 = ruta_anexo4

    def procesar_datos(self) -> tuple[Gestor_Proyectos, Gestor_ProyectosConcedidos, Gestor_ProyectosContrato]:
        gestor_general = Gestor_Proyectos()
        gestor_concedidos = Gestor_ProyectosConcedidos()
        gestor_contratos = Gestor_ProyectosContrato()

        self._leer_denegados(gestor_general)
        self._leer_concedidos(gestor_general, gestor_concedidos)
        self._leer_contratos(gestor_general, gestor_concedidos, gestor_contratos)

        return gestor_general, gestor_concedidos, gestor_contratos

    def _parsear_float(self, valor) -> float:
        """
        Convierte el valor de una celda de Excel a float de forma segura.
        Soporta celdas vacías (None), números nativos y textos con comas.
        """
        if valor is None or str(valor).strip() == "":
            return 0.0
        
        # Si Excel ya lo lee como número (int o float)
        if isinstance(valor, (int, float)):
            return float(valor)
            
        v = str(valor).strip()
        # Limpieza de textos con formato de moneda español
        if '.' in v and ',' in v:
            v = v.replace('.', '')
        v = v.replace(',', '.')
        
        try:
            return float(v)
        except ValueError:
            return 0.0

    def _leer_denegados(self, gestor_general: Gestor_Proyectos):
        """Lee el Anexo III (Denegados) y crea objetos Proyecto."""
        wb = openpyxl.load_workbook(self.ruta_anexo3, data_only=True)
        ws = wb.active
        
        # iter_rows con min_row=2 se salta automáticamente la cabecera
        for fila in ws.iter_rows(min_row=2, values_only=True):
            # Si la fila está vacía o no tiene Referencia en la columna 2 (índice 1)
            if not fila or fila[1] is None: 
                continue
                
            proyecto_denegado = Proyecto(
                referencia=str(fila[1]).strip(),
                area=str(fila[2]).strip() if fila[2] else "",
                entidad_solicitante=str(fila[3]).strip() if fila[3] else "",
                comunidad_autonoma=str(fila[5]).strip() if fila[5] else ""
            )
            gestor_general.agregar_proyecto(proyecto_denegado)
        wb.close()

    def _leer_concedidos(self, gestor_general: Gestor_Proyectos, gestor_concedidos: Gestor_ProyectosConcedidos):
        """Cruza datos de Anexo I y Anexo II para crear ProyectoConcedido."""
        datos_base = {}
        
        # 1. Leer datos base del Anexo I
        wb1 = openpyxl.load_workbook(self.ruta_anexo1, data_only=True)
        ws1 = wb1.active
        for fila in ws1.iter_rows(min_row=2, values_only=True):
            if not fila or fila[1] is None: 
                continue
            referencia = str(fila[1]).strip()
            datos_base[referencia] = {
                'area': str(fila[2]).strip() if fila[2] else "",
                'entidad': str(fila[3]).strip() if fila[3] else "",
                'ccaa': str(fila[5]).strip() if fila[5] else ""
            }
        wb1.close()

        # 2. Leer importes del Anexo II y cruzar
        wb2 = openpyxl.load_workbook(self.ruta_anexo2, data_only=True)
        ws2 = wb2.active
        for fila in ws2.iter_rows(min_row=2, values_only=True):
            if not fila or fila[1] is None: 
                continue
                
            referencia = str(fila[1]).strip()

            if referencia in datos_base:
                base = datos_base[referencia]
                
                proyecto_concedido = ProyectoConcedido(
                    referencia=referencia,
                    area=base['area'],
                    entidad_solicitante=base['entidad'],
                    comunidad_autonoma=base['ccaa'],
                    costes_directos=self._parsear_float(fila[3]),
                    costes_indirectos=self._parsear_float(fila[4]),
                    anticipo=self._parsear_float(fila[5]),
                    subvencion=self._parsear_float(fila[6]),
                    anualidades=[
                        self._parsear_float(fila[8]),
                        self._parsear_float(fila[9]),
                        self._parsear_float(fila[10]),
                        self._parsear_float(fila[11])
                    ],
                    num_contratos=int(self._parsear_float(fila[12]))
                )
                gestor_general.agregar_proyecto(proyecto_concedido)
                gestor_concedidos.agregar_proyecto(proyecto_concedido)
        wb2.close()

    def _leer_contratos(self, gestor_general: Gestor_Proyectos, gestor_concedidos: Gestor_ProyectosConcedidos, gestor_contratos: Gestor_ProyectosContrato):
        """Lee el Anexo IV y actualiza a ProyectoContrato."""
        wb4 = openpyxl.load_workbook(self.ruta_anexo4, data_only=True)
        ws4 = wb4.active
        for fila in ws4.iter_rows(min_row=2, values_only=True):
            if not fila or fila[5] is None: 
                continue
                
            titulo = str(fila[4]).strip() if fila[4] else ""
            referencia = str(fila[5]).strip()

            if referencia in gestor_concedidos.proyectos_concedidos:
                proy_orig = gestor_concedidos.proyectos_concedidos[referencia]
                
                proyecto_contrato = ProyectoContrato(
                    referencia=proy_orig.referencia,
                    area=proy_orig.area,
                    entidad_solicitante=proy_orig.entidad_solicitante,
                    comunidad_autonoma=proy_orig.comunidad_autonoma,
                    costes_directos=proy_orig.costes_directos,
                    costes_indirectos=proy_orig.costes_indirectos,
                    anticipo=proy_orig.anticipo,
                    subvencion=proy_orig.subvencion,
                    anualidades=proy_orig.anualidades,
                    num_contratos=1,
                    titulo=titulo
                )
                
                gestor_general.agregar_proyecto(proyecto_contrato)
                gestor_concedidos.agregar_proyecto(proyecto_contrato)
                gestor_contratos.agregar_proyecto(proyecto_contrato)
        wb4.close()
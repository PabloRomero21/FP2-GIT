import openpyxl
from nomenclator import Nomenclator

class FactoriaNomenclator:
    """
    Clase Factoría encargada de leer el archivo .xlsx y construir el Nomenclator.
    """

    @staticmethod
    def construir_nomenclator(ruta_excel):
        """
        Abre el Excel, lee las dos primeras hojas y devuelve el Nomenclator lleno.
        """
        nomenclator = Nomenclator()
        
        # Cargar el libro de Excel (data_only=True ignora fórmulas y trae solo los valores reales)
        libro = openpyxl.load_workbook(ruta_excel, data_only=True)
        hojas = libro.sheetnames
        
        # El enunciado especifica: primera hoja Hombres, segunda Mujeres
        if len(hojas) >= 2:
            FactoriaNomenclator._leer_hoja(libro[hojas[0]], True, nomenclator)   # Hombres
            FactoriaNomenclator._leer_hoja(libro[hojas[1]], False, nomenclator)  # Mujeres
            
        return nomenclator

    @staticmethod
    def _leer_hoja(hoja, es_hombre, nomenclator):
        """
        Procesa una hoja concreta del Excel buscando las décadas y sus datos.
        """
        indices_decadas = []
        
        # 1. RASTREO DE DÉCADAS (Buscamos la palabra "NACID" en la primera fila)
        for col in range(1, hoja.max_column + 1):
            valor_celda = hoja.cell(row=1, column=col).value
            if valor_celda is not None:
                texto_valor = str(valor_celda).strip().upper()
                
                # Buscamos "NACID" para que atrape tanto "NACIDOS" como "NACIDAS"
                if "NACID" in texto_valor:
                    # Limpiamos el texto para que quede más bonito en la terminal
                    # Ej: "NACIDOS EN AÑOS 1930 A 1939" -> "1930 A 1939"
                    decada_limpia = texto_valor.replace("NACIDOS EN AÑOS ", "").replace("NACIDAS EN AÑOS ", "")
                    decada_limpia = decada_limpia.replace("NACIDOS ", "").replace("NACIDAS ", "")
                    
                    # Guardamos la columna base donde empieza esta década y su nombre
                    indices_decadas.append((col, decada_limpia))

        # 2. EXTRACCIÓN DE DATOS
        # Los datos de nombres empiezan en la fila 3
        for fila in range(3, hoja.max_row + 1):
            
            for col_base, decada in indices_decadas:
                # En el Excel del INE, debajo de la celda de la década tenemos:
                # Misma columna (col_base) = Nombre
                # Columna siguiente (col_base + 1) = Frecuencia
                # Siguiente (col_base + 2) = Tanto por mil
                
                celda_nombre = hoja.cell(row=fila, column=col_base).value
                celda_frec = hoja.cell(row=fila, column=col_base + 1).value
                celda_tpm = hoja.cell(row=fila, column=col_base + 2).value
                
                # Verificamos que la celda tenga un nombre válido y no sea la cabecera "NOMBRE"
                if celda_nombre and str(celda_nombre).strip() != "" and str(celda_nombre).upper() != "NOMBRE":
                    try:
                        nombre_limpio = str(celda_nombre).strip()
                        
                        # Limpieza de números (por si el Excel los lee como texto con puntos/comas)
                        str_frec = str(celda_frec).replace('.', '').replace(',', '')
                        str_tpm = str(celda_tpm).replace(',', '.')
                        
                        # Convertimos a números (usamos int(float()) por si viene algo como '1200.0')
                        frecuencia_abs = int(float(str_frec)) 
                        tanto_por_mil = float(str_tpm)
                        
                        # Recuperamos o creamos el objeto Nombre en el Nomenclator
                        objeto_nombre = nomenclator.obtener_nombre(nombre_limpio, es_hombre)
                        
                        # Añadimos los datos de esta década
                        objeto_nombre.añadir_datos_decada(decada, frecuencia_abs, tanto_por_mil)
                        
                    except (ValueError, TypeError):
                        # Ignoramos celdas corruptas, vacías o que no se puedan convertir a números
                        pass